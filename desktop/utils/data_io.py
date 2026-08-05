# -*- coding: utf-8 -*-
"""
数据导入导出工具模块
支持 CSV / Excel 格式的导出与导入，以及全量 JSON 备份/恢复
"""
import csv
import json
import sqlite3
import sys
from datetime import datetime
from PyQt5.QtWidgets import QFileDialog, QMessageBox
from database.db_manager import get_connection, _safe_sql_identifier

from utils.nutstore_sync import get_sync as _get_sync
def _sync_cloud():
    try:
        _get_sync().trigger_sync()
    except Exception as e:
        import logging; logging.getLogger(__name__).debug(f"云同步失败: {e}")

# ===== 允许操作的表名白名单（防止 SQL 注入） =====
_ALLOWED_TABLES = frozenset([
    "employees", "finance_records", "daily_revenue", "purchases",
    "ingredients", "suppliers", "attendance", "salary_records",
    "reimbursements", "stores", "departments", "approvals",
    "leave_records",
])

def _validate_table_name(table):
    """验证表名是否在白名单中，防止 SQL 注入"""
    if table not in _ALLOWED_TABLES:
        raise ValueError(f"非法表名: {table}")
    return table

# ===== 表名 -> 中文标题映射 =====
TABLE_LABELS = {
    "employees": "员工数据",
    "finance_records": "收支记录",
    "daily_revenue": "营业额记录",
    "purchases": "进货记录",
    "ingredients": "食材库存",
    "suppliers": "供应商",
    "attendance": "考勤记录",
    "salary_records": "工资记录",
    "reimbursements": "报销记录",
}

# ===== 各表导出列配置 (table -> [(db_column, display_header), ...]) =====
EXPORT_COLUMNS = {
    "employees": [
        ("name", "姓名"), ("phone", "手机号"), ("position", "职位"),
        ("base_salary", "基本工资"), ("hire_date", "入职日期"),
        ("status", "状态"), ("role", "角色"), ("remark", "备注"),
    ],
    "finance_records": [
        ("record_date", "日期"), ("record_type", "类型"), ("category", "类别"),
        ("amount", "金额"), ("account", "支付方式"), ("operator", "经办人"),
        ("description", "说明"), ("remark", "备注"),
    ],
    "daily_revenue": [
        ("record_date", "日期"), ("channel", "渠道"),
        ("package_name", "套餐"), ("package_type", "类型"),
        ("order_count", "数量"), ("amount", "金额"),
        ("remark", "备注"),
    ],
    "ingredients": [
        ("name", "名称"), ("category", "分类"), ("unit", "单位"),
        ("price", "单价"), ("stock", "当前库存"), ("min_stock", "最低库存"),
        ("remark", "备注"),
    ],
    "suppliers": [
        ("name", "名称"), ("contact", "联系人"), ("phone", "电话"),
        ("address", "地址"), ("remark", "备注"),
    ],
}

def _get_table_data(table, store_filter=None):
    """从数据库读取表数据，支持门店过滤"""
    _validate_table_name(table)
    conn = get_connection()
    cursor = conn.cursor()
    config = EXPORT_COLUMNS.get(table)
    if not config:
        conn.close()
        return [], []

    cols = [c[0] for c in config]
    headers = [c[1] for c in config]
    col_str = ", ".join(cols)

    safe_table = _safe_sql_identifier(table)
    if store_filter and "store_id" in _get_table_columns(cursor, table):
        cursor.execute(
            f"SELECT {col_str} FROM {safe_table} WHERE store_id=? OR store_id IS NULL ORDER BY id",
            (store_filter,)
        )
    else:
        cursor.execute(f"SELECT {col_str} FROM {safe_table} ORDER BY id")

    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return headers, rows

def _get_table_columns(cursor, table):
    """获取表的所有列名"""
    cursor.execute(f"PRAGMA table_info({_safe_sql_identifier(table)})")
    return {row[1] for row in cursor.fetchall()}

def _import_openpyxl():
    """安全导入 openpyxl
    
    openpyxl 3.x 的可选依赖 numpy 在某些版本组合下存在兼容性问题。
    此函数在导入 openpyxl 前临时标记 numpy 为不可用，导入后恢复。
    """
    import importlib
    _saved = sys.modules.get('numpy')
    _saved_sub = {}
    for key in list(sys.modules.keys()):
        if key.startswith('numpy.'):
            _saved_sub[key] = sys.modules.pop(key)
    # 设置 numpy 为 None，让 openpyxl 的 import numpy 触发 ImportError 并被捕获
    sys.modules['numpy'] = None
    try:
        import openpyxl
        return openpyxl
    except Exception:
        raise
    finally:
        # 恢复 numpy 状态
        sys.modules.pop('numpy', None)
        if _saved is not None:
            sys.modules['numpy'] = _saved
        sys.modules.update(_saved_sub)

def export_data_to_excel(data, headers, filename):
    """将原始数据列表导出为 Excel 文件（供 cost_calc / salary 调用）
    
    Args:
        data: 二维列表，每行是一条记录
        headers: 表头列表
        filename: 默认文件名
    Returns:
        保存路径 or None
    """
    try:
        openpyxl = _import_openpyxl()
    except Exception:
        # openpyxl 不可用，回退到 CSV
        import csv
        path, _ = QFileDialog.getSaveFileName(
            None, "导出 CSV", filename.replace('.xlsx', '.csv'), "CSV 文件 (*.csv)")
        if not path:
            return None
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            writer.writerows(data)
        return path

    path, _ = QFileDialog.getSaveFileName(
        None, "导出 Excel", filename, "Excel 文件 (*.xlsx)")
    if not path:
        return None
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据"
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        # 自动列宽
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, len(data) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)
        wb.save(path)
        return path
    except Exception as e:
        QMessageBox.warning(None, "导出失败", f"导出出错：{e}")
        return None

def export_to_csv(table, parent_widget=None, store_filter=None):
    """导出指定表数据为 CSV 文件"""
    headers, rows = _get_table_data(table, store_filter)
    if not rows:
        if parent_widget:
            QMessageBox.information(parent_widget, "提示", "没有可导出的数据")
        return False

    label = TABLE_LABELS.get(table, table)
    default_name = f"{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    path, _ = QFileDialog.getSaveFileName(
        parent_widget, "导出 CSV", default_name, "CSV 文件 (*.csv)"
    )
    if not path:
        return False

    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row.get(c[0], "") for c in EXPORT_COLUMNS[table]])
        if parent_widget:
            QMessageBox.information(parent_widget, "导出成功", f"已导出 {len(rows)} 条记录到:\n{path}")
        return True
    except Exception as e:
        if parent_widget:
            QMessageBox.warning(parent_widget, "导出失败", f"导出出错：{e}")
        return False

def export_to_excel(table, parent_widget=None, store_filter=None):
    """导出指定表数据为 Excel 文件"""
    # 尝试导入 openpyxl，失败则回退到 CSV
    try:
        openpyxl = _import_openpyxl()
    except Exception as e:
        try:
            from utils.logger import logger
            logger.warning(f"openpyxl 导入失败，回退到CSV: {e}")
        except Exception:
            pass
        return export_to_csv(table, parent_widget, store_filter)

    headers, rows = _get_table_data(table, store_filter)
    if not rows:
        if parent_widget:
            QMessageBox.information(parent_widget, "提示", "没有可导出的数据")
        return False

    label = TABLE_LABELS.get(table, table)
    default_name = f"{label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path, _ = QFileDialog.getSaveFileName(
        parent_widget, "导出 Excel", default_name, "Excel 文件 (*.xlsx)"
    )
    if not path:
        return False

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = label[:31]
        # 写表头
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
        # 写数据
        for row_idx, row in enumerate(rows, 2):
            for col_idx, col_config in enumerate(EXPORT_COLUMNS[table], 1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_config[0], ""))
        # 自动列宽
        for col_idx in range(1, len(headers) + 1):
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for row_idx in range(2, len(rows) + 2):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)

        wb.save(path)
        if parent_widget:
            QMessageBox.information(parent_widget, "导出成功", f"已导出 {len(rows)} 条记录到:\n{path}")
        return True
    except Exception as e:
        if parent_widget:
            QMessageBox.warning(parent_widget, "导出失败", f"导出出错：{e}")
        return False

def import_from_csv(table, parent_widget=None):
    """从 CSV 文件导入数据，返回 (成功条数, 跳过条数)"""
    _validate_table_name(table)
    label = TABLE_LABELS.get(table, table)
    path, _ = QFileDialog.getOpenFileName(
        parent_widget, f"导入 {label}", "", "CSV 文件 (*.csv)"
    )
    if not path:
        return 0, 0

    config = EXPORT_COLUMNS.get(table)
    if not config:
        return 0, 0

    try:
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            file_headers = next(reader, None)
            if not file_headers:
                return 0, 0

            # 建立文件列名 -> 数据库列名 的映射
            header_map = {}
            for i, fh in enumerate(file_headers):
                fh_clean = fh.strip()
                for db_col, display_header in config:
                    if fh_clean == display_header:
                        header_map[i] = db_col
                        break

            if not header_map:
                if parent_widget:
                    QMessageBox.warning(parent_widget, "格式错误",
                        f"CSV 列头不匹配。\n需要: {[c[1] for c in config]}\n实际: {file_headers}")
                return 0, 0

            conn = get_connection()
            cursor = conn.cursor()
            db_cols = [c[0] for c in config]
            # 获取表的所有列
            all_cols = _get_table_columns(cursor, table)
            # 只导入数据库中存在的列
            insert_cols = [db_col for db_col in db_cols if db_col in all_cols and db_col != "id"]
            placeholders = ", ".join(["?"] * len(insert_cols))
            col_str = ", ".join(insert_cols)

            success = 0
            skipped = 0
            for row in reader:
                if not row or all(not cell.strip() for cell in row):
                    skipped += 1
                    continue
                data = {}
                for i, val in enumerate(row):
                    if i in header_map:
                        data[header_map[i]] = val.strip()

            # 检查必填字段
                if table == "employees" and not data.get("name"):
                    skipped += 1
                    continue
                if table == "ingredients" and not data.get("name"):
                    skipped += 1
                    continue

                values = []
                for col in insert_cols:
                    val = data.get(col, "")
            # 尝试数值转换
                    if col in ("base_salary", "price", "stock", "min_stock", "amount"):
                        try:
                            val = float(val) if val else 0
                        except ValueError:
                            val = 0
                    elif col == "order_count":
                        try:
                            val = int(val) if val else 0
                        except ValueError:
                            val = 0
                    values.append(val)

                try:
                    cursor.execute(
                        f"INSERT INTO {table} ({col_str}) VALUES ({placeholders})",
                        values
                    )
                    success += 1
                except sqlite3.IntegrityError:
                    skipped += 1
                except Exception:
                    skipped += 1

            conn.commit()
            _sync_cloud()

            if parent_widget:
                QMessageBox.information(parent_widget, "导入完成",
                    f"成功导入 {success} 条记录\n跳过 {skipped} 条（重复或格式错误）")
            return success, skipped
    except Exception as e:
        if parent_widget:
            QMessageBox.warning(parent_widget, "导入失败", f"导入出错：{e}")
        return 0, 0

def full_backup(parent_widget=None):
    """全量备份所有业务表为 JSON 文件"""
    tables = list(EXPORT_COLUMNS.keys()) + ["stores", "departments", "approvals", "leave_records"]
    backup_data = {}
    conn = get_connection()
    cursor = conn.cursor()
    for table in tables:
        _validate_table_name(table)
        try:
            cursor.execute(f"SELECT * FROM {_safe_sql_identifier(table)}")
            rows = [dict(r) for r in cursor.fetchall()]
            backup_data[table] = rows
        except Exception as e:
            import logging; logging.getLogger(__name__).debug(f"数据操作失败(可忽略): {e}")
    conn.close()

    default_name = f"full_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    path, _ = QFileDialog.getSaveFileName(
        parent_widget, "全量备份", default_name, "JSON 备份 (*.json)"
    )
    if not path:
        return False

    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(backup_data, f, ensure_ascii=False, indent=2)
        if parent_widget:
            total = sum(len(v) for v in backup_data.values())
            QMessageBox.information(parent_widget, "备份成功",
                f"已备份 {len(backup_data)} 张表共 {total} 条记录到:\n{path}")
        return True
    except Exception as e:
        if parent_widget:
            QMessageBox.warning(parent_widget, "备份失败", f"备份出错：{e}")
        return False

def full_restore(parent_widget=None):
    """从 JSON 备份文件恢复数据"""
    path, _ = QFileDialog.getOpenFileName(
        parent_widget, "恢复数据", "", "JSON 备份 (*.json)"
    )
    if not path:
        return False

    try:
        with open(path, "r", encoding="utf-8") as f:
            backup_data = json.load(f)
    except Exception as e:
        if parent_widget:
            QMessageBox.warning(parent_widget, "恢复失败", f"文件读取失败：{e}")
        return False

    reply = QMessageBox.question(
        parent_widget, "确认恢复",
        "恢复将覆盖当前所有数据，此操作不可撤销！\n"
        "注意：若目标表存在唯一约束，冲突行将被跳过，可能导致数据不完整。\n\n"
        "确定要继续吗？",
        QMessageBox.Yes | QMessageBox.No, QMessageBox.No
    )
    if reply != QMessageBox.Yes:
        return False

    conn = get_connection()
    cursor = conn.cursor()
    restored = 0
    for table, rows in backup_data.items():
        try:
            _validate_table_name(table)
    # 清空现有数据
            cursor.execute(f"DELETE FROM {_safe_sql_identifier(table)}")
    # 逐条插入
            for row in rows:
                cols = list(row.keys())
                vals = list(row.values())
                placeholders = ", ".join(["?"] * len(cols))
                col_str = ", ".join(cols)
                try:
                    cursor.execute(f"INSERT OR IGNORE INTO {_safe_sql_identifier(table)} ({col_str}) VALUES ({placeholders})", vals)
                    if cursor.rowcount > 0:
                        restored += 1
                except Exception as e:
                    import logging; logging.getLogger(__name__).debug(f"数据操作失败(可忽略): {e}")
        except Exception as e:
            import logging; logging.getLogger(__name__).debug(f"数据操作失败(可忽略): {e}")

    conn.commit()
    conn.close()
    _sync_cloud()
    if parent_widget:
        QMessageBox.information(parent_widget, "恢复成功", f"已恢复 {restored} 条记录")
    return True
